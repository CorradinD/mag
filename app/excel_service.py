from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import Settings


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


@dataclass
class ExcelRowResult:
    file_path: str
    status: str
    row_number: int | None = None
    sku: str | None = None
    message: str | None = None
    old_price: float | None = None
    new_price: float | None = None


def ensure_allowed_file(path: Path, offers_dir: Path) -> Path:
    resolved = path.resolve()
    root = offers_dir.resolve()
    if resolved.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Formato non supportato: {resolved.suffix}")
    if root not in resolved.parents and resolved != root:
        raise ValueError("Il file deve trovarsi nella cartella offerte configurata")
    if not resolved.exists():
        raise FileNotFoundError(str(resolved))
    return resolved


def backup_file(path: Path, backup_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination_dir = backup_dir / timestamp
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / path.name
    shutil.copy2(path, destination)
    return destination


def update_offer_file(path: Path, prices: dict[str, float], settings: Settings) -> list[ExcelRowResult]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return update_xlsx_offer(path, prices, settings)
    if suffix == ".xls":
        return update_xls_offer(path, prices, settings)
    raise ValueError(f"Formato non supportato: {suffix}")


def update_xlsx_offer(path: Path, prices: dict[str, float], settings: Settings) -> list[ExcelRowResult]:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, keep_vba=keep_vba)
    try:
        if settings.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Foglio non trovato: {settings.sheet_name}")

        sheet = workbook[settings.sheet_name]
        results: list[ExcelRowResult] = []
        changed = False
        for row in range(settings.data_start_row, sheet.max_row + 1):
            sku = sheet[f"{settings.sku_column}{row}"].value
            normalized_sku = str(sku).strip() if sku is not None else ""
            if not normalized_sku:
                results.append(ExcelRowResult(str(path), "missing_sku", row, None, "SKU mancante"))
                continue
            if normalized_sku not in prices:
                results.append(ExcelRowResult(str(path), "unknown_sku", row, normalized_sku, "SKU non presente in app"))
                continue

            price_cell = sheet[f"{settings.price_column}{row}"]
            old_price = _to_float(price_cell.value)
            new_price = prices[normalized_sku]
            if old_price != new_price:
                price_cell.value = new_price
                changed = True
            results.append(ExcelRowResult(str(path), "updated", row, normalized_sku, None, old_price, new_price))

        if changed:
            workbook.save(path)
        return results
    finally:
        workbook.close()


def update_xls_offer(path: Path, prices: dict[str, float], settings: Settings) -> list[ExcelRowResult]:
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Per aggiornare file .xls serve Windows con pywin32 e Microsoft Excel installato") from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
        if workbook.ReadOnly:
            raise RuntimeError("File .xls aperto in sola lettura o bloccato da un altro utente")
        sheet = workbook.Worksheets(settings.sheet_name)
        last_row = sheet.UsedRange.Rows.Count
        results: list[ExcelRowResult] = []
        for row in range(settings.data_start_row, last_row + 1):
            sku = sheet.Range(f"{settings.sku_column}{row}").Value
            normalized_sku = str(sku).strip() if sku is not None else ""
            if normalized_sku.endswith(".0"):
                normalized_sku = normalized_sku[:-2]
            if not normalized_sku:
                results.append(ExcelRowResult(str(path), "missing_sku", row, None, "SKU mancante"))
                continue
            if normalized_sku not in prices:
                results.append(ExcelRowResult(str(path), "unknown_sku", row, normalized_sku, "SKU non presente in app"))
                continue

            cell = sheet.Range(f"{settings.price_column}{row}")
            old_price = _to_float(cell.Value)
            new_price = prices[normalized_sku]
            if old_price != new_price:
                cell.Value = new_price
            results.append(ExcelRowResult(str(path), "updated", row, normalized_sku, None, old_price, new_price))
        workbook.Save()
        return results
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
