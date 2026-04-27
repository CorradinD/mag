from __future__ import annotations

import importlib
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


@pytest.fixture()
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    monkeypatch.setenv("MAG_DATABASE", str(tmp_path / "magazzino.sqlite3"))
    monkeypatch.setenv("MAG_OFFERS_DIR", str(tmp_path / "offerte"))
    monkeypatch.setenv("MAG_BACKUP_DIR", str(tmp_path / "backup"))
    monkeypatch.setenv("MAG_EXCEL_SHEET", "Offerta")
    monkeypatch.setenv("MAG_EXCEL_SKU_COLUMN", "A")
    monkeypatch.setenv("MAG_EXCEL_PRICE_COLUMN", "D")

    import app.main as main

    importlib.reload(main)
    with TestClient(main.app) as test_client:
        yield test_client


def test_create_item_and_prevent_duplicate_sku(client: TestClient) -> None:
    response = client.post(
        "/items",
        data={"sku": "ABC123", "description": "Profilo alluminio", "price": "12.50", "quantity": "3"},
    )
    assert response.status_code == 200
    assert response.json()["sku"] == "ABC123"

    duplicate = client.post(
        "/items",
        data={"sku": "ABC123", "description": "Duplicato", "price": "10", "quantity": "0"},
    )
    assert duplicate.status_code == 409


def test_movements_update_quantity_and_block_negative_stock(client: TestClient) -> None:
    client.post("/items", data={"sku": "MAT1", "description": "Materiale", "price": "5", "quantity": "2"})

    load = client.post(
        "/movements",
        data={"sku": "MAT1", "movement_type": "carico", "quantity": "4", "operator": "Mario"},
    )
    assert load.status_code == 200

    unload = client.post("/movements", data={"sku": "MAT1", "movement_type": "scarico", "quantity": "3"})
    assert unload.status_code == 200

    items = client.get("/items").json()
    assert items[0]["quantity"] == 3

    too_much = client.post("/movements", data={"sku": "MAT1", "movement_type": "scarico", "quantity": "4"})
    assert too_much.status_code == 400
    assert "Giacenza insufficiente" in too_much.json()["detail"]


def test_update_xlsx_offer_and_create_backup(client: TestClient, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    offers_dir = tmp_path / "offerte"
    backup_dir = tmp_path / "backup"
    offer_path = offers_dir / "offerta.xlsx"

    client.post("/items", data={"sku": "SKU1", "description": "Articolo 1", "price": "99.90", "quantity": "1"})
    client.post("/items", data={"sku": "SKU2", "description": "Articolo 2", "price": "42", "quantity": "1"})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Offerta"
    sheet["A1"] = "SKU"
    sheet["D1"] = "Prezzo"
    sheet["A2"] = "SKU1"
    sheet["D2"] = 10
    sheet["A3"] = "MISSING"
    sheet["D3"] = 20
    sheet["A4"] = None
    sheet["D4"] = 30
    workbook.save(offer_path)

    response = client.post("/excel/update-offers", data={"files": ["offerta.xlsx"]})
    assert response.status_code == 200
    result = response.json()
    assert result["rows_updated"] == 1
    assert result["rows_unknown_sku"] == 1
    assert result["rows_missing_sku"] == 1
    assert result["files_updated"] == 1

    updated = load_workbook(offer_path)
    assert updated["Offerta"]["D2"].value == 99.9
    updated.close()

    backups = list(backup_dir.glob("*/offerta.xlsx"))
    assert len(backups) == 1

